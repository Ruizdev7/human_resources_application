import os
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from flask import Blueprint, make_response, send_file

from hhrr_app.models.tbl_employee import Employee
from hhrr_app.models.tbl_ss_employee import SSEmployee
from hhrr_app.models.tbl_family_nucleus import FamilyNucleus
from hhrr_app.models.tbl_health_condition import HealthCondition
from hhrr_app.models.tbl_demographic_data import DemographicData
from hhrr_app.models.tbl_sociodemographic_data import SociodemographicData
from hhrr_app.models.tbl_employment_relationship import EmploymentRelationship
from hhrr_app.models.tbl_emergency_contact_details import EmergencyContactDetails


excel_bp = Blueprint("excel", __name__)


@excel_bp.route("api/v1/generate-employee-excel-data", methods=["GET"])
def generar_excel_employees():
    employees = Employee.query.all()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Empleados"

    # Add excel headers
    sheet["A1"] = "CCN"
    sheet["B1"] = "Tipo de Documento"
    sheet["C1"] = "Numero de Identificación"
    sheet["D1"] = "Nombre Completo"
    sheet["E1"] = "Fecha de Nacimiento"
    sheet["F1"] = "Edad"
    sheet["G1"] = "Rango de Edad"
    sheet["H1"] = "Genero"
    sheet["I1"] = "RH"
    sheet["J1"] = "E-mail Personal"
    sheet["K1"] = "Telefono Corporativo"
    sheet["L1"] = "Estado Civil"

    row = 2
    for employee in employees:
        sheet.cell(row=row, column=1, value=employee.ccn_employee)
        sheet.cell(row=row, column=2, value=employee.TypeId.description_type_id)
        sheet.cell(row=row, column=3, value=employee.number_id_employee)
        sheet.cell(row=row, column=4, value=employee.full_name_employee)
        sheet.cell(row=row, column=5, value=employee.date_birth_employee)
        sheet.cell(row=row, column=6, value=employee.age)
        sheet.cell(row=row, column=7, value=employee.AgeRange.age_range)
        sheet.cell(
            row=row, column=8, value=employee.AutoPerceivedGender.auto_perceived_gender
        )
        sheet.cell(row=row, column=9, value=employee.RH.rh)
        sheet.cell(row=row, column=10, value=employee.employee_personal_email)
        sheet.cell(row=row, column=11, value=employee.employee_personal_cellphone)
        sheet.cell(row=row, column=12, value=employee.MaritalStatus.marital_status)
        row += 1

    temp_file = NamedTemporaryFile(delete=False)
    temp_filename = temp_file.name
    wb.save(temp_filename)
    wb.close()

    response = make_response(send_file(temp_filename, as_attachment=True))
    response.headers["Content-Disposition"] = "attachment; filename=empleados.xlsx"

    os.unlink(temp_filename)

    return response


@excel_bp.route("api/v1/generate-employment-relationship-excel-data", methods=["GET"])
def generar_excel_employment_relation():
    employees = EmploymentRelationship.query.all()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Vinculación Laboral"

    # Add excel headers
    sheet["A1"] = "CCN"
    sheet["B1"] = "Nombre del Empleado"
    sheet["C1"] = "Rol"
    sheet["D1"] = "Area"
    sheet["E1"] = "Proceso"
    sheet["F1"] = "Turno de Trabajo"
    sheet["G1"] = "Fecha de Vinculación"
    sheet["H1"] = "Tiempo Trabajado"
    sheet["I1"] = "Tipo de Vinculación"
    sheet["J1"] = "E-mail Corporativo"
    sheet["K1"] = "Telefono Corporativo"
    sheet["L1"] = "Jefe Inmediato"
    sheet["M1"] = "Manager"
    sheet["N1"] = "Tipo de Cargo"
    sheet["O1"] = "Es Empleado Activo"

    row = 2
    for employee in employees:
        sheet.cell(row=row, column=1, value=employee.ccn_employee)
        sheet.cell(row=row, column=2, value=employee.Employee.full_name_employee)
        sheet.cell(row=row, column=3, value=employee.Role.role)
        sheet.cell(row=row, column=4, value=employee.Role.area)
        sheet.cell(row=row, column=5, value=employee.Role.process)
        sheet.cell(row=row, column=6, value=employee.WorkShift.description_work_shift)
        sheet.cell(row=row, column=7, value=employee.binding_date)
        sheet.cell(row=row, column=8, value=employee.time_worked)
        sheet.cell(
            row=row,
            column=9,
            value=employee.TypeRelationship.description_type_relationship,
        )
        sheet.cell(row=row, column=10, value=employee.employee_corporate_email)
        sheet.cell(row=row, column=11, value=employee.employee_corporate_cellphone)
        sheet.cell(row=row, column=12, value=employee.ImmediateBoss.full_name_employee)
        sheet.cell(row=row, column=13, value=employee.Manager.full_name_employee)
        sheet.cell(row=row, column=14, value=employee.type_of_charge)
        sheet.cell(row=row, column=15, value=employee.is_active_employee)
        row += 1

    temp_file = NamedTemporaryFile(delete=False)
    temp_filename = temp_file.name
    wb.save(temp_filename)
    wb.close()

    response = make_response(send_file(temp_filename, as_attachment=True))
    response.headers[
        "Content-Disposition"
    ] = "attachment; filename=vinculacon_laboral.xlsx"

    os.unlink(temp_filename)

    return response


@excel_bp.route("api/v1/generate-family-nucleus-excel-data", methods=["GET"])
def generar_excel_family_nucleus():
    employees = FamilyNucleus.query.all()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Nucleo Familiar"

    # Add excel headers
    sheet["A1"] = "CCN"
    sheet["B1"] = "Nombre Completo"
    sheet["C1"] = "No. de Hijos"
    sheet["D1"] = "Tipo de Documento"
    sheet["E1"] = "Tipo de Documento"
    sheet["F1"] = "No. Identificación"
    sheet["G1"] = "Genero"
    sheet["H1"] = "Primer Nombre"
    sheet["I1"] = "Segundo Nombre"
    sheet["J1"] = "Primer Apellido"
    sheet["K1"] = "Segundo Apellido"
    sheet["L1"] = "Fecha de Nacimiento"
    sheet["M1"] = "Edad"
    sheet["N1"] = "Rango de Edad"
    sheet["O1"] = "Nivel de Escolaridad"

    row = 2
    for employee in employees:
        sheet.cell(row=row, column=1, value=employee.ccn_employee)
        sheet.cell(row=row, column=2, value=employee.Employee.full_name_employee)
        sheet.cell(row=row, column=3, value=employee.number_of_children)
        sheet.cell(row=row, column=4, value=employee.TypeId.type_id)
        sheet.cell(row=row, column=5, value=employee.TypeId.description_type_id)
        sheet.cell(row=row, column=6, value=employee.number_id)
        sheet.cell(
            row=row, column=7, value=employee.AutoPerceivedGender.auto_perceived_gender
        )
        sheet.cell(row=row, column=8, value=employee.first_name)
        sheet.cell(row=row, column=9, value=employee.middle_name)
        sheet.cell(row=row, column=10, value=employee.first_last_name)
        sheet.cell(row=row, column=11, value=employee.second_last_name)
        sheet.cell(row=row, column=12, value=employee.date_of_birth)
        sheet.cell(row=row, column=13, value=employee.age)
        sheet.cell(row=row, column=14, value=employee.AgeRange.age_range)
        sheet.cell(
            row=row,
            column=15,
            value=employee.SchoolingLevel.description_schooling_level,
        )
        row += 1

    temp_file = NamedTemporaryFile(delete=False)
    temp_filename = temp_file.name
    wb.save(temp_filename)
    wb.close()

    response = make_response(send_file(temp_filename, as_attachment=True))
    response.headers[
        "Content-Disposition"
    ] = "attachment; filename=nucleo_familiar.xlsx"

    os.unlink(temp_filename)

    return response


@excel_bp.route("api/v1/generate-demographic-data-excel-data", methods=["GET"])
def generar_excel_demographic_data():
    employees = DemographicData.query.all()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Datos Demograficos"

    # Add excel headers
    sheet["A1"] = "CCN"
    sheet["B1"] = "Nombre Completo"
    sheet["C1"] = "Departamento de Nacimiento"
    sheet["D1"] = "Ciudad de Nacimiento"
    sheet["E1"] = "Departamento de Recidencia"
    sheet["F1"] = "Ciudad de Recidencia"
    sheet["G1"] = "Nivel de Escolaridad"
    sheet["H1"] = "Raza"
    sheet["I1"] = "Es cabeza de Hogar"

    row = 2
    for employee in employees:
        sheet.cell(row=row, column=1, value=employee.ccn_employee)
        sheet.cell(row=row, column=2, value=employee.Employee.full_name_employee)
        sheet.cell(
            row=row, column=3, value=employee.DepartmentBirth.descripcion_department
        )
        sheet.cell(row=row, column=4, value=employee.CityBirthCity.description_city)
        sheet.cell(
            row=row, column=5, value=employee.DepartmentResidence.descripcion_department
        )
        sheet.cell(row=row, column=6, value=employee.CityCityResidence.description_city)
        sheet.cell(
            row=row, column=7, value=employee.SchoolingLevel.description_schooling_level
        )
        sheet.cell(row=row, column=8, value=employee.Race.description_race)
        sheet.cell(row=row, column=9, value=employee.is_head_of_household)

        row += 1

    temp_file = NamedTemporaryFile(delete=False)
    temp_filename = temp_file.name
    wb.save(temp_filename)
    wb.close()

    response = make_response(send_file(temp_filename, as_attachment=True))
    response.headers[
        "Content-Disposition"
    ] = "attachment; filename=datos_demograficos.xlsx"

    os.unlink(temp_filename)

    return response


@excel_bp.route("api/v1/generate-ss-employee-excel-data", methods=["GET"])
def generar_excel_ss_employee():
    employees = SSEmployee.query.all()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Seguridad Social"

    # Add excel headers
    sheet["A1"] = "CCN"
    sheet["B1"] = "Nombre Completo"
    sheet["C1"] = "Tipo de Afiliación"
    sheet["D1"] = "Tipo de Contribuidor"
    sheet["E1"] = "EPS"
    sheet["F1"] = "AFP"
    sheet["G1"] = "ARL"
    sheet["H1"] = "CCF"
    sheet["I1"] = "AAP"

    row = 2
    for employee in employees:
        sheet.cell(row=row, column=1, value=employee.ccn_employee)
        sheet.cell(row=row, column=2, value=employee.Employee.full_name_employee)
        sheet.cell(
            row=row,
            column=3,
            value=employee.TypeAffiliation.description_type_affiliation,
        )
        sheet.cell(
            row=row,
            column=4,
            value=employee.TypeContributor.description_type_contributor,
        )
        sheet.cell(row=row, column=5, value=employee.EPS.description_eps)
        sheet.cell(row=row, column=6, value=employee.AFP.description_afp)
        sheet.cell(row=row, column=7, value=employee.ARL.description_arl)
        sheet.cell(row=row, column=8, value=employee.CCF.description_ccf)
        sheet.cell(row=row, column=9, value=employee.AAP.description_aap)

        row += 1

    temp_file = NamedTemporaryFile(delete=False)
    temp_filename = temp_file.name
    wb.save(temp_filename)
    wb.close()

    response = make_response(send_file(temp_filename, as_attachment=True))
    response.headers[
        "Content-Disposition"
    ] = "attachment; filename=seguridad_social.xlsx"

    os.unlink(temp_filename)

    return response


@excel_bp.route("api/v1/generate-health-condition-excel-data", methods=["GET"])
def generar_excel_health_condition():
    employees = HealthCondition.query.all()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Condicion de Salud"

    # Add excel headers
    sheet["A1"] = "CCN"
    sheet["B1"] = "Nombre Completo"
    sheet["C1"] = "consume bebidas alcoholicas?"
    sheet["D1"] = "Enfermedades"
    sheet["E1"] = "Alergias"
    sheet["F1"] = "Cual(es)?"
    sheet["G1"] = "Medicamentos"
    sheet["H1"] = "Cual(es)?"
    sheet["I1"] = "Ultima Consulta Medica"
    sheet["J1"] = "A pensado ingerir menos bebidas alcoholicas"
    sheet["K1"] = "Le molesta las criticas por ingerir alcohol"
    sheet["L1"] = "Siente necesidad por ingerir alcohol en horas de la mañana"
    sheet["M1"] = "actividad fisica 3 veces por semana minimo 30 minutos"
    sheet["N1"] = "fumador"
    sheet["O1"] = "Cuantos cigarrillos al dia"
    sheet["P1"] = "es exfumador"
    sheet["Q1"] = "consume sustancias psicoactivas"
    sheet["R1"] = "antes consumia sustancias psicoactivas"
    sheet["S1"] = "cual sustancia psicoactiva"
    row = 2
    for employee in employees:
        sheet.cell(row=row, column=1, value=employee.ccn_employee)
        sheet.cell(row=row, column=2, value=employee.Employee.full_name_employee)
        sheet.cell(row=row, column=3, value=employee.consume_alcoholic_beverages)
        sheet.cell(row=row, column=4, value=employee.Diseases.diseases)
        sheet.cell(row=row, column=5, value=employee.allergies)
        sheet.cell(row=row, column=6, value=employee.what_allergy)
        sheet.cell(row=row, column=7, value=employee.medicines)
        sheet.cell(row=row, column=8, value=employee.what_medicin)
        sheet.cell(row=row, column=9, value=employee.last_medical_consultation)
        sheet.cell(
            row=row, column=10, value=employee.plan_to_drink_less_alcoholic_beverages
        )
        sheet.cell(
            row=row,
            column=11,
            value=employee.discomfort_due_to_criticism_when_ingesting_alcohol,
        )
        sheet.cell(
            row=row, column=13, value=employee.need_to_drink_alcohol_in_the_morning
        )
        sheet.cell(
            row=row,
            column=14,
            value=employee.physical_activity_3_times_a_week_30_minutes,
        )
        sheet.cell(row=row, column=15, value=employee.he_is_a_smoker)
        sheet.cell(row=row, column=16, value=employee.how_many_cigarettes_a_day)
        sheet.cell(row=row, column=17, value=employee.he_is_ex_smoker)
        sheet.cell(row=row, column=18, value=employee.consume_psychoactive_substances)
        sheet.cell(
            row=row, column=19, value=employee.used_psychoactive_substances_before
        )
        sheet.cell(row=row, column=20, value=employee.what_psychoactive_substances)
        row += 1

    temp_file = NamedTemporaryFile(delete=False)
    temp_filename = temp_file.name
    wb.save(temp_filename)
    wb.close()

    response = make_response(send_file(temp_filename, as_attachment=True))
    response.headers[
        "Content-Disposition"
    ] = "attachment; filename=condicion_de_salud.xlsx"

    os.unlink(temp_filename)

    return response


@excel_bp.route("api/v1/generate-datos-sociodemograficos-excel-data", methods=["GET"])
def generar_excel_datos_sociodemograficos():
    employees = SociodemographicData.query.all()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Datos Sociodemograficos"

    # Add excel headers
    sheet["A1"] = "CCN"
    sheet["B1"] = "Nombre Completo"
    sheet["C1"] = "Personas que Dependen"
    sheet["D1"] = "Personas en el Hogar"
    sheet["E1"] = "Personas con Discapacidad en el Hogar"
    sheet["F1"] = "Ingresos Familiares"
    sheet["G1"] = "Los Ingresos Alcanzan"
    sheet["H1"] = "Caracteristicas de Vivienda"
    sheet["I1"] = "Tipo de Vivienda"
    sheet["J1"] = "Localización"
    sheet["K1"] = "Dirección de Residencia"
    sheet["L1"] = "Transporte para ir a Trabajar"
    sheet["M1"] = "Transporte Opcional para ir a Trabajar"
    sheet["N1"] = "Estrato"
    sheet["O1"] = "Energia Electrica"
    sheet["P1"] = "Alcantarillado"
    sheet["Q1"] = "Acueducto"
    sheet["R1"] = "Gas Natural"
    sheet["S1"] = "Recolección de Basuras"
    sheet["T1"] = "Telefono Personal"
    sheet["U1"] = "Tiene Deudas"
    sheet["V1"] = "Desea Refinanciar las Deudas"
    sheet["W1"] = "Tiene Computador en Casa"
    sheet["X1"] = "Tiene Internet en Casa"

    row = 2
    for employee in employees:
        sheet.cell(row=row, column=1, value=employee.ccn_employee)
        sheet.cell(row=row, column=2, value=employee.Employee.full_name_employee)
        sheet.cell(row=row, column=3, value=employee.other_dependents)
        sheet.cell(row=row, column=4, value=employee.relatives)
        sheet.cell(row=row, column=5, value=employee.people_with_disabilities)
        sheet.cell(row=row, column=6, value=employee.monthly_income)
        sheet.cell(row=row, column=7, value=employee.is_income_enougth)
        sheet.cell(row=row, column=8, value=employee.SubHouseType.sub_house_type)
        sheet.cell(row=row, column=9, value=employee.HouseType.house_type)
        sheet.cell(row=row, column=10, value=employee.where_its_located)
        sheet.cell(row=row, column=11, value=employee.residence_address)
        sheet.cell(row=row, column=12, value=employee.type_transportation)
        sheet.cell(row=row, column=13, value=employee.type_transportation_2)
        sheet.cell(row=row, column=14, value=employee.social_stratum)
        sheet.cell(row=row, column=15, value=employee.electric_power)
        sheet.cell(row=row, column=16, value=employee.sewerage)
        sheet.cell(row=row, column=17, value=employee.aqueduct)
        sheet.cell(row=row, column=18, value=employee.natural_gas)
        sheet.cell(row=row, column=19, value=employee.garbage_collection)
        sheet.cell(row=row, column=20, value=employee.landline)
        sheet.cell(row=row, column=21, value=employee.debts)
        sheet.cell(row=row, column=22, value=employee.debt_refinancing)
        sheet.cell(row=row, column=23, value=employee.computer_at_home)
        sheet.cell(row=row, column=24, value=employee.have_internet_access)

        row += 1

    temp_file = NamedTemporaryFile(delete=False)
    temp_filename = temp_file.name
    wb.save(temp_filename)
    wb.close()

    response = make_response(send_file(temp_filename, as_attachment=True))
    response.headers[
        "Content-Disposition"
    ] = "attachment; filename=datos_sociodemograficos.xlsx"

    os.unlink(temp_filename)

    return response


@excel_bp.route("api/v1/generate-emergency-contact-detail-excel-data", methods=["GET"])
def generar_excel_emergency_contact_detail():
    employees = EmergencyContactDetails.query.all()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Contacto de Emergencia"

    # Add excel headers
    sheet["A1"] = "CCN"
    sheet["B1"] = "Nombre Completo"
    sheet["C1"] = "Contacto de Emergencia"
    sheet["D1"] = "Parentezco"
    sheet["E1"] = "Telefono"

    row = 2
    for employee in employees:
        sheet.cell(row=row, column=1, value=employee.ccn_employee)
        sheet.cell(row=row, column=2, value=employee.Employee.full_name_employee)
        sheet.cell(row=row, column=3, value=employee.emergency_contact)
        sheet.cell(row=row, column=4, value=employee.ccn_relationship)
        sheet.cell(row=row, column=5, value=employee.cellphone)
        row += 1

    temp_file = NamedTemporaryFile(delete=False)
    temp_filename = temp_file.name
    wb.save(temp_filename)
    wb.close()

    response = make_response(send_file(temp_filename, as_attachment=True))
    response.headers[
        "Content-Disposition"
    ] = "attachment; filename=contacto_de_emergencia.xlsx"

    os.unlink(temp_filename)

    return response
